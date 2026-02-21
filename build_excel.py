"""
ClearMetric FIRE Calculator Pro — Premium Excel Template
Product for Gumroad ($14.99)

4 Sheets:
  1. FIRE Calculator — inputs, FIRE number, key metrics, action items
  2. Scenario Comparison — compare 3 different retirement scenarios side by side
  3. Year-by-Year Projection — 40-year detailed breakdown
  4. How To Use — instructions and FIRE education

Design: Emerald Green palette (growth/wealth theme)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
import os

# ============================================================
# DESIGN SYSTEM — Emerald Green (growth/wealth)
# ============================================================
EMERALD = "1A7A4C"
DARK_EMERALD = "0D5E35"
WHITE = "FFFFFF"
INPUT_GREEN = "D5F5E3"
LIGHT_GRAY = "F5F6FA"
MED_GRAY = "D5D8DC"
DARK_GRAY = "5D6D7E"
GREEN = "27AE60"
LIGHT_GREEN = "EAFAF1"
RED = "E74C3C"
LIGHT_RED = "FDEDEC"
YELLOW = "F39C12"
LIGHT_YELLOW = "FEF9E7"
ACCENT = "2ECC71"
GOLD = "F0B90B"
LIGHT_EMERALD = "E8F8F0"

FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=12, color="A9DFBF", italic=True)
FONT_SECTION = Font(name="Calibri", size=13, bold=True, color=WHITE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_LABEL = Font(name="Calibri", size=11, color="2C3E50")
FONT_INPUT = Font(name="Calibri", size=12, color=EMERALD, bold=True)
FONT_VALUE = Font(name="Calibri", size=11, color="2C3E50")
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color=EMERALD)
FONT_SMALL = Font(name="Calibri", size=9, color=DARK_GRAY, italic=True)
FONT_BIG = Font(name="Calibri", size=28, bold=True, color=WHITE)
FONT_BIG_LABEL = Font(name="Calibri", size=12, bold=True, color="A9DFBF")
FONT_GREEN = Font(name="Calibri", size=11, bold=True, color=GREEN)
FONT_RED = Font(name="Calibri", size=11, bold=True, color=RED)
FONT_YELLOW = Font(name="Calibri", size=11, bold=True, color=YELLOW)
FONT_LINK = Font(name="Calibri", size=11, color=ACCENT, underline="single")
FONT_CTA = Font(name="Calibri", size=12, bold=True, color=EMERALD)
FONT_WHITE_BOLD = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_FIRE = Font(name="Calibri", size=36, bold=True, color=GOLD)
FONT_FIRE_LABEL = Font(name="Calibri", size=11, bold=True, color=WHITE)

FILL_EMERALD = PatternFill(start_color=EMERALD, end_color=EMERALD, fill_type="solid")
FILL_DARK = PatternFill(start_color=DARK_EMERALD, end_color=DARK_EMERALD, fill_type="solid")
FILL_INPUT = PatternFill(start_color=INPUT_GREEN, end_color=INPUT_GREEN, fill_type="solid")
FILL_GRAY = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
FILL_LIGHT = PatternFill(start_color=LIGHT_EMERALD, end_color=LIGHT_EMERALD, fill_type="solid")
FILL_GREEN = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
FILL_RED = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
FILL_YELLOW = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
FILL_ACCENT = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
FILL_GOLD = PatternFill(start_color="FEF5D4", end_color="FEF5D4", fill_type="solid")

THIN = Border(
    left=Side("thin", MED_GRAY), right=Side("thin", MED_GRAY),
    top=Side("thin", MED_GRAY), bottom=Side("thin", MED_GRAY),
)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center")


def header_bar(ws, row, c1, c2, text, fill=None):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=text)
    cell.font = FONT_SECTION
    cell.fill = fill or FILL_EMERALD
    cell.alignment = ALIGN_C
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = fill or FILL_EMERALD
        ws.cell(row=row, column=c).border = THIN


def label_input(ws, row, lc, vc, label, value=None, fmt=None, hint=None):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=value)
    cv.font = FONT_INPUT
    cv.fill = FILL_INPUT
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt
    if hint:
        # Place hint one column after input
        ch = ws.cell(row=row, column=vc + 1, value=hint)
        ch.font = FONT_SMALL
        ch.alignment = ALIGN_L


def label_calc(ws, row, lc, vc, label, formula, fmt=None, bold=False):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=formula)
    cv.font = FONT_BOLD if bold else FONT_VALUE
    cv.fill = FILL_WHITE
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt


def cols(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


# ============================================================
# SHEET 1: FIRE CALCULATOR
# ============================================================
def build_fire_calculator(ws):
    ws.title = "FIRE Calculator"
    ws.sheet_properties.tabColor = EMERALD
    cols(ws, {
        "A": 2, "B": 36, "C": 18, "D": 18, "E": 4,
        "F": 36, "G": 18, "H": 2,
    })

    # Background
    for r in range(1, 75):
        for c in range(1, 9):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    # ===== TITLE =====
    for r in range(1, 4):
        for c in range(2, 8):
            ws.cell(row=r, column=c).fill = FILL_DARK

    ws.merge_cells("B1:G1")
    ws.row_dimensions[1].height = 10

    ws.merge_cells("B2:G2")
    ws.row_dimensions[2].height = 38
    title = ws.cell(row=2, column=2, value="FIRE CALCULATOR PRO")
    title.font = FONT_TITLE
    title.alignment = ALIGN_C

    ws.merge_cells("B3:G3")
    ws.row_dimensions[3].height = 22
    sub = ws.cell(row=3, column=2,
                  value="Financial Independence, Retire Early — Enter your numbers. Get your plan.")
    sub.font = FONT_SUBTITLE
    sub.alignment = ALIGN_C

    # ===== LEFT COLUMN: INPUTS =====
    # --- Personal Info ---
    header_bar(ws, 5, 2, 4, "YOUR PROFILE")

    label_input(ws, 6, 2, 3, "Current Age", 30, "0", "years")
    label_input(ws, 7, 2, 3, "Target Retirement Age", 45, "0", "FIRE target")
    label_input(ws, 8, 2, 3, "Life Expectancy", 90, "0", "plan conservatively")

    # --- Income & Expenses ---
    header_bar(ws, 10, 2, 4, "INCOME & EXPENSES")

    label_input(ws, 11, 2, 3, "Annual Gross Income", 100000, "$#,##0")
    label_input(ws, 12, 2, 3, "Effective Tax Rate", 0.25, "0.0%")
    label_input(ws, 13, 2, 3, "Monthly Living Expenses", 4000, "$#,##0")
    label_input(ws, 14, 2, 3, "Annual Income Growth Rate", 0.03, "0.0%", "raises, promotions")

    # Calculated
    label_calc(ws, 15, 2, 3, "After-Tax Income (annual)",
               "=C11*(1-C12)", "$#,##0", bold=True)
    label_calc(ws, 16, 2, 3, "Annual Living Expenses",
               "=C13*12", "$#,##0", bold=True)

    # --- Savings & Investments ---
    header_bar(ws, 18, 2, 4, "SAVINGS & INVESTMENTS")

    label_input(ws, 19, 2, 3, "Current Invested Assets", 150000, "$#,##0")
    label_input(ws, 20, 2, 3, "Monthly Investment Amount", 2500, "$#,##0")
    label_input(ws, 21, 2, 3, "Stock Allocation", 0.70, "0%", "e.g. 70%")
    label_input(ws, 22, 2, 3, "Bond/Fixed Allocation", "=1-C21", "0%")

    # --- Assumptions ---
    header_bar(ws, 24, 2, 4, "MARKET ASSUMPTIONS")

    label_input(ws, 25, 2, 3, "Expected Stock Return (nominal)", 0.10, "0.0%", "long-term ~10%")
    label_input(ws, 26, 2, 3, "Expected Bond Return (nominal)", 0.05, "0.0%", "long-term ~5%")
    label_input(ws, 27, 2, 3, "Expected Inflation Rate", 0.03, "0.0%", "US avg ~3%")
    label_input(ws, 28, 2, 3, "Safe Withdrawal Rate (SWR)", 0.04, "0.0%", "4% rule")

    # --- Social Security ---
    header_bar(ws, 30, 2, 4, "SOCIAL SECURITY / PENSION")

    label_input(ws, 31, 2, 3, "Monthly SS/Pension Amount", 0, "$#,##0", "$0 if none")
    label_input(ws, 32, 2, 3, "SS/Pension Starting Age", 67, "0")

    # ===== RIGHT COLUMN: RESULTS =====

    # --- FIRE Number (big display) ---
    header_bar(ws, 5, 6, 7, "YOUR FIRE NUMBER", FILL_DARK)

    ws.merge_cells("F6:G9")
    for r in range(6, 10):
        for c in range(6, 8):
            ws.cell(row=r, column=c).fill = FILL_DARK
            ws.cell(row=r, column=c).border = THIN

    fire_cell = ws.cell(row=6, column=6)
    fire_cell.value = '=IF(C28>0,(C13*12*(1+C27)^(C7-C6))/C28,"Set SWR")'
    fire_cell.font = FONT_FIRE
    fire_cell.alignment = ALIGN_C
    fire_cell.number_format = "$#,##0"

    ws.merge_cells("F10:G10")
    for c in range(6, 8):
        ws.cell(row=10, column=c).fill = FILL_DARK
        ws.cell(row=10, column=c).border = THIN
    desc = ws.cell(row=10, column=6, value="The portfolio needed to cover your expenses forever")
    desc.font = Font(name="Calibri", size=10, italic=True, color="A9DFBF")
    desc.alignment = ALIGN_C

    # --- Key Metrics ---
    header_bar(ws, 12, 6, 7, "KEY METRICS")

    # Blended nominal return
    label_calc(ws, 13, 6, 7, "Blended Nominal Return",
               "=C21*C25+(1-C21)*C26", "0.0%")
    # Real return
    label_calc(ws, 14, 6, 7, "Real Return (after inflation)",
               "=G13-C27", "0.0%")
    # Years to FIRE
    label_calc(ws, 15, 6, 7, "Years to FIRE",
               "=C7-C6", "0")
    # Annual investment
    label_calc(ws, 16, 6, 7, "Annual Investment",
               "=C20*12", "$#,##0")
    # Savings Rate
    label_calc(ws, 17, 6, 7, "Savings Rate",
               '=IF(C15>0,C20*12/C15,0)', "0.0%", bold=True)
    # Monthly expenses at retirement
    label_calc(ws, 18, 6, 7, "Monthly Expenses at Retirement",
               "=C13*(1+C27)^(C7-C6)", "$#,##0")
    # Annual expenses at retirement
    label_calc(ws, 19, 6, 7, "Annual Expenses at Retirement",
               "=G18*12", "$#,##0")

    # --- Projected Corpus ---
    header_bar(ws, 21, 6, 7, "PROJECTED CORPUS AT RETIREMENT")

    # FV of current savings (simplified for Excel: compound annually)
    label_calc(ws, 22, 6, 7, "Growth of Current Savings",
               "=C19*(1+G14)^G15", "$#,##0")
    # FV of monthly investments (annuity with growth — simplified annual)
    # Using annuity formula: PMT * ((1+r)^n - 1) / r
    label_calc(ws, 23, 6, 7, "Value of Future Investments",
               '=IF(G14=0,C20*12*G15,C20*12*((1+G14)^G15-1)/G14)', "$#,##0")
    # Total projected corpus
    label_calc(ws, 24, 6, 7, "TOTAL PROJECTED CORPUS",
               "=G22+G23", "$#,##0", bold=True)
    ws.cell(row=24, column=7).font = Font(name="Calibri", size=13, bold=True, color=EMERALD)

    # Gap / Surplus
    label_calc(ws, 25, 6, 7, "Gap / Surplus",
               "=G24-F6", "$#,##0", bold=True)

    # Status
    label_calc(ws, 26, 6, 7, "Status",
               '=IF(G25>=0,"ON TRACK ✅","SHORTFALL ❌")', None, bold=True)

    # Conditional formatting for gap
    ws.conditional_formatting.add("G25",
        CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                   fill=FILL_GREEN, font=FONT_GREEN))
    ws.conditional_formatting.add("G25",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=FILL_RED, font=FONT_RED))

    # --- Action Items ---
    header_bar(ws, 28, 6, 7, "ACTION ITEMS")

    # Additional monthly investment needed
    label_calc(ws, 29, 6, 7, "Extra Monthly Investment Needed",
               '=IF(G25>=0,0,IF(G14=0,-G25/(G15*12),'
               '-G25*((1+G14)^(1/12)-1)/((1+G14)^G15-1)/12))',
               "$#,##0", bold=True)

    # Coast FIRE number
    label_calc(ws, 30, 6, 7, "Coast FIRE Number",
               "=F6/(1+G14)^G15", "$#,##0")
    ws.cell(row=30, column=6).font = FONT_SMALL
    # Coast FIRE status
    label_calc(ws, 31, 6, 7, "Coast FIRE Status",
               '=IF(C19>=G30,"ACHIEVED — You could stop investing! ✅",'
               '"Need $"&TEXT(G30-C19,"#,##0")&" more to coast")', None)

    # Barista FIRE (partial income covers expenses, investments grow)
    label_calc(ws, 32, 6, 7, "Barista FIRE Income Needed",
               "=MAX(G19-C19*C28,0)/12", "$#,##0")

    # --- Savings Rate Context ---
    header_bar(ws, 34, 6, 7, "SAVINGS RATE → YEARS TO FI")

    savings_data = [
        (35, "10% savings rate", "~51 years"),
        (36, "15% savings rate", "~43 years"),
        (37, "25% savings rate", "~32 years"),
        (38, "30% savings rate", "~28 years"),
        (39, "40% savings rate", "~22 years"),
        (40, "50% savings rate", "~17 years"),
        (41, "60% savings rate", "~12 years"),
        (42, "70% savings rate", "~8 years"),
        (43, "80% savings rate", "~5 years"),
    ]
    for r, label, years in savings_data:
        ws.row_dimensions[r].height = 18
        ws.cell(row=r, column=6, value=label).font = FONT_LABEL
        ws.cell(row=r, column=6).fill = FILL_GRAY
        ws.cell(row=r, column=6).border = THIN
        ws.cell(row=r, column=7, value=years).font = FONT_VALUE
        ws.cell(row=r, column=7).fill = FILL_GRAY
        ws.cell(row=r, column=7).border = THIN
        ws.cell(row=r, column=7).alignment = ALIGN_C

    # Highlight user's savings rate row
    for r, label, _ in savings_data:
        pct_str = label.split("%")[0].strip()
        pct_val = int(pct_str) / 100
        next_idx = savings_data.index((r, label, _))
        if next_idx < len(savings_data) - 1:
            next_pct = int(savings_data[next_idx + 1][1].split("%")[0].strip()) / 100
        else:
            next_pct = 1.0
        ws.conditional_formatting.add(
            f"F{r}:G{r}",
            FormulaRule(formula=[f"AND($G$17>={pct_val},$G$17<{next_pct})"],
                        fill=FILL_GOLD,
                        font=Font(name="Calibri", size=11, bold=True, color=EMERALD)))

    # --- CTA ---
    row = 46
    ws.merge_cells(f"B{row}:G{row}")
    ws.cell(row=row, column=2,
            value="© ClearMetric | clearmetric.gumroad.com | Professional FIRE Planning").font = FONT_SMALL
    ws.cell(row=row, column=2).alignment = ALIGN_C

    # Protection
    ws.protection.sheet = True
    ws.protection.set_password("")
    input_cells = [
        (6, 3), (7, 3), (8, 3),
        (11, 3), (12, 3), (13, 3), (14, 3),
        (19, 3), (20, 3), (21, 3),
        (25, 3), (26, 3), (27, 3), (28, 3),
        (31, 3), (32, 3),
    ]
    for r, c in input_cells:
        ws.cell(row=r, column=c).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 2: SCENARIO COMPARISON
# ============================================================
def build_scenario_comparison(wb):
    ws = wb.create_sheet("Scenario Comparison")
    ws.sheet_properties.tabColor = "2E86C1"
    cols(ws, {
        "A": 2, "B": 34, "C": 18, "D": 4, "E": 18, "F": 4, "G": 18, "H": 2,
    })

    for r in range(1, 55):
        for c in range(1, 9):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    # Title
    for r in range(1, 4):
        for c in range(2, 8):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:G1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:G2")
    ws.row_dimensions[2].height = 38
    ws.cell(row=2, column=2, value="SCENARIO COMPARISON").font = FONT_TITLE
    ws.cell(row=2, column=2).alignment = ALIGN_C
    ws.merge_cells("B3:G3")
    ws.cell(row=3, column=2,
            value="Compare 3 different paths to FIRE. Green cells = your inputs.").font = FONT_SUBTITLE
    ws.cell(row=3, column=2).alignment = ALIGN_C

    # Column headers
    for c in range(2, 8):
        ws.cell(row=5, column=c).fill = FILL_EMERALD
        ws.cell(row=5, column=c).border = THIN
    ws.cell(row=5, column=2, value="Parameter").font = FONT_WHITE_BOLD
    ws.cell(row=5, column=2).alignment = ALIGN_C
    ws.cell(row=5, column=3, value="Current Plan").font = FONT_WHITE_BOLD
    ws.cell(row=5, column=3).alignment = ALIGN_C
    ws.cell(row=5, column=5, value="Aggressive").font = FONT_WHITE_BOLD
    ws.cell(row=5, column=5).alignment = ALIGN_C
    ws.cell(row=5, column=7, value="Conservative").font = FONT_WHITE_BOLD
    ws.cell(row=5, column=7).alignment = ALIGN_C

    # Pull from main sheet for Scenario A, manual for B & C
    params = [
        (6, "Current Age", "='FIRE Calculator'!C6", 30, 30, "0"),
        (7, "Target Retirement Age", "='FIRE Calculator'!C7", 40, 55, "0"),
        (8, "Life Expectancy", "='FIRE Calculator'!C8", 90, 90, "0"),
        (9, "Monthly Expenses", "='FIRE Calculator'!C13", 3000, 5000, "$#,##0"),
        (10, "Current Savings", "='FIRE Calculator'!C19", 150000, 150000, "$#,##0"),
        (11, "Monthly Investment", "='FIRE Calculator'!C20", 4000, 1500, "$#,##0"),
        (12, "Stock Allocation", "='FIRE Calculator'!C21", 0.90, 0.50, "0%"),
        (13, "Inflation Rate", "='FIRE Calculator'!C27", 0.03, 0.03, "0.0%"),
        (14, "Safe Withdrawal Rate", "='FIRE Calculator'!C28", 0.04, 0.035, "0.0%"),
    ]

    for r, label, val_a, val_b, val_c, fmt in params:
        ws.row_dimensions[r].height = 25
        ws.cell(row=r, column=2, value=label).font = FONT_LABEL
        ws.cell(row=r, column=2).fill = FILL_GRAY
        ws.cell(row=r, column=2).border = THIN

        # Scenario A (linked to main sheet)
        ca = ws.cell(row=r, column=3, value=val_a)
        ca.font = FONT_BOLD
        ca.fill = FILL_LIGHT
        ca.border = THIN
        ca.alignment = ALIGN_C
        if fmt:
            ca.number_format = fmt

        # Scenario B
        cb = ws.cell(row=r, column=5, value=val_b)
        cb.font = FONT_INPUT
        cb.fill = FILL_INPUT
        cb.border = THIN
        cb.alignment = ALIGN_C
        if fmt:
            cb.number_format = fmt

        # Scenario C
        cc = ws.cell(row=r, column=7, value=val_c)
        cc.font = FONT_INPUT
        cc.fill = FILL_INPUT
        cc.border = THIN
        cc.alignment = ALIGN_C
        if fmt:
            cc.number_format = fmt

    # Calculated results
    header_bar(ws, 16, 2, 7, "RESULTS")

    def _result_row(r, label, formula_a, formula_b, formula_c, fmt):
        ws.row_dimensions[r].height = 25
        ws.cell(row=r, column=2, value=label).font = FONT_LABEL
        ws.cell(row=r, column=2).fill = FILL_GRAY
        ws.cell(row=r, column=2).border = THIN
        for col, formula in [(3, formula_a), (5, formula_b), (7, formula_c)]:
            cell = ws.cell(row=r, column=col, value=formula)
            cell.font = FONT_BOLD
            cell.fill = FILL_WHITE
            cell.border = THIN
            cell.alignment = ALIGN_C
            if fmt:
                cell.number_format = fmt

    # Blended return (for each scenario)
    _result_row(17, "Nominal Return",
                "=C12*0.10+(1-C12)*0.05", "=E12*0.10+(1-E12)*0.05", "=G12*0.10+(1-G12)*0.05", "0.0%")
    _result_row(18, "Real Return",
                "=C17-C13", "=E17-E13", "=G17-G13", "0.0%")
    _result_row(19, "Years to FIRE",
                "=C7-C6", "=E7-E6", "=G7-G6", "0")
    _result_row(20, "FIRE Number",
                "=IF(C14>0,(C9*12*(1+C13)^C19)/C14,0)",
                "=IF(E14>0,(E9*12*(1+E13)^E19)/E14,0)",
                "=IF(G14>0,(G9*12*(1+G13)^G19)/G14,0)",
                "$#,##0")
    _result_row(21, "Projected Corpus",
                "=C10*(1+C18)^C19+IF(C18=0,C11*12*C19,C11*12*((1+C18)^C19-1)/C18)",
                "=E10*(1+E18)^E19+IF(E18=0,E11*12*E19,E11*12*((1+E18)^E19-1)/E18)",
                "=G10*(1+G18)^G19+IF(G18=0,G11*12*G19,G11*12*((1+G18)^G19-1)/G18)",
                "$#,##0")
    _result_row(22, "Gap / Surplus",
                "=C21-C20", "=E21-E20", "=G21-G20", "$#,##0")
    _result_row(23, "Status",
                '=IF(C22>=0,"ON TRACK ✅","SHORTFALL ❌")',
                '=IF(E22>=0,"ON TRACK ✅","SHORTFALL ❌")',
                '=IF(G22>=0,"ON TRACK ✅","SHORTFALL ❌")',
                None)
    _result_row(24, "Savings Rate",
                "=IF(C6>0,C11*12/(100000*(1-0.25)),0)",
                "=IF(E6>0,E11*12/(100000*(1-0.25)),0)",
                "=IF(G6>0,G11*12/(100000*(1-0.25)),0)",
                "0.0%")

    # Conditional formatting for gap rows
    for col_letter in ["C", "E", "G"]:
        ws.conditional_formatting.add(
            f"{col_letter}22",
            CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                       fill=FILL_GREEN, font=FONT_GREEN))
        ws.conditional_formatting.add(
            f"{col_letter}22",
            CellIsRule(operator="lessThan", formula=["0"],
                       fill=FILL_RED, font=FONT_RED))

    # Winner
    ws.merge_cells("B26:G26")
    ws.cell(row=26, column=2,
            value='=IF(AND(C22>=E22,C22>=G22),"→ Current Plan is the best path",'
                  'IF(AND(E22>=C22,E22>=G22),"→ Aggressive scenario wins",'
                  '"→ Conservative scenario wins"))').font = FONT_CTA
    ws.cell(row=26, column=2).alignment = ALIGN_C

    # Protect calculated cells
    ws.protection.sheet = True
    ws.protection.set_password("")
    for r in range(6, 15):
        for c in [5, 7]:
            ws.cell(row=r, column=c).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 3: YEAR-BY-YEAR PROJECTION
# ============================================================
def build_yearly_projection(wb):
    ws = wb.create_sheet("Year-by-Year")
    ws.sheet_properties.tabColor = GREEN
    cols(ws, {
        "A": 2, "B": 6, "C": 6, "D": 16, "E": 16,
        "F": 16, "G": 16, "H": 16, "I": 10, "J": 2,
    })

    for r in range(1, 4):
        for c in range(2, 10):
            ws.cell(row=r, column=c).fill = FILL_DARK

    ws.merge_cells("B1:I1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:I2")
    ws.row_dimensions[2].height = 38
    ws.cell(row=2, column=2, value="YEAR-BY-YEAR FIRE PROJECTION").font = FONT_TITLE
    ws.cell(row=2, column=2).alignment = ALIGN_C
    ws.merge_cells("B3:I3")
    ws.cell(row=3, column=2,
            value="All values in real (inflation-adjusted) terms. Based on your inputs in the FIRE Calculator sheet.").font = FONT_SUBTITLE
    ws.cell(row=3, column=2).alignment = ALIGN_C

    # Headers
    headers = ["Year", "Age", "Portfolio Start", "Annual Investment",
               "Investment Growth", "Portfolio End", "FIRE Number", "% of FIRE"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=5, column=2 + i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_EMERALD
        cell.alignment = ALIGN_C
        cell.border = THIN

    fc = "'FIRE Calculator'"

    # 40 years of projection
    for yr in range(41):
        r = 6 + yr
        ws.row_dimensions[r].height = 20

        # Year
        ws.cell(row=r, column=2, value=yr).font = FONT_VALUE
        ws.cell(row=r, column=2).alignment = ALIGN_C
        ws.cell(row=r, column=2).border = THIN
        ws.cell(row=r, column=2).fill = FILL_GRAY

        # Age
        ws.cell(row=r, column=3, value=f"={fc}!C6+{yr}").font = FONT_VALUE
        ws.cell(row=r, column=3).alignment = ALIGN_C
        ws.cell(row=r, column=3).border = THIN
        ws.cell(row=r, column=3).fill = FILL_GRAY

        if yr == 0:
            # Portfolio start = current savings
            ws.cell(row=r, column=4, value=f"={fc}!C19").font = FONT_VALUE
            ws.cell(row=r, column=4).number_format = "$#,##0"
            ws.cell(row=r, column=4).border = THIN

            # Annual investment
            ws.cell(row=r, column=5, value=f"={fc}!C20*12").font = FONT_VALUE
            ws.cell(row=r, column=5).number_format = "$#,##0"
            ws.cell(row=r, column=5).border = THIN

            # Investment growth
            ws.cell(row=r, column=6, value=f"=D{r}*{fc}!G14").font = FONT_VALUE
            ws.cell(row=r, column=6).number_format = "$#,##0"
            ws.cell(row=r, column=6).border = THIN

            # Portfolio end
            ws.cell(row=r, column=7, value=f"=D{r}+E{r}+F{r}").font = FONT_BOLD
            ws.cell(row=r, column=7).number_format = "$#,##0"
            ws.cell(row=r, column=7).border = THIN
        else:
            # Portfolio start = previous end
            ws.cell(row=r, column=4, value=f"=G{r-1}").font = FONT_VALUE
            ws.cell(row=r, column=4).number_format = "$#,##0"
            ws.cell(row=r, column=4).border = THIN

            # Annual investment (grows with income growth)
            ws.cell(row=r, column=5,
                    value=f"=E{r-1}*(1+{fc}!C14)").font = FONT_VALUE
            ws.cell(row=r, column=5).number_format = "$#,##0"
            ws.cell(row=r, column=5).border = THIN

            # Investment growth
            ws.cell(row=r, column=6, value=f"=D{r}*{fc}!G14").font = FONT_VALUE
            ws.cell(row=r, column=6).number_format = "$#,##0"
            ws.cell(row=r, column=6).border = THIN

            # Portfolio end
            ws.cell(row=r, column=7, value=f"=D{r}+E{r}+F{r}").font = FONT_BOLD
            ws.cell(row=r, column=7).number_format = "$#,##0"
            ws.cell(row=r, column=7).border = THIN

        # FIRE Number (inflated expenses / SWR — stays same in real terms)
        ws.cell(row=r, column=8,
                value=f"={fc}!C13*12/{fc}!C28").font = FONT_VALUE
        ws.cell(row=r, column=8).number_format = "$#,##0"
        ws.cell(row=r, column=8).border = THIN
        ws.cell(row=r, column=8).fill = FILL_LIGHT

        # % of FIRE
        ws.cell(row=r, column=9,
                value=f"=IF(H{r}>0,G{r}/H{r},0)").font = FONT_BOLD
        ws.cell(row=r, column=9).number_format = "0.0%"
        ws.cell(row=r, column=9).border = THIN

        # Color the row alternately
        if yr % 2 == 1:
            for c in range(4, 10):
                ws.cell(row=r, column=c).fill = FILL_GRAY

        # Highlight the retirement year
        ws.conditional_formatting.add(
            f"B{r}:I{r}",
            FormulaRule(formula=[f"C{r}={fc}!C7"],
                        fill=FILL_GOLD,
                        font=Font(name="Calibri", size=11, bold=True, color=EMERALD)))

        # Highlight when FIRE achieved
        ws.conditional_formatting.add(
            f"I{r}",
            CellIsRule(operator="greaterThanOrEqual", formula=["1"],
                       fill=FILL_GREEN, font=FONT_GREEN))

    # Net worth growth chart
    chart = LineChart()
    chart.title = "Portfolio Growth vs FIRE Number"
    chart.style = 10
    chart.y_axis.title = "Portfolio Value ($)"
    chart.y_axis.numFmt = "$#,##0"
    chart.x_axis.title = "Age"

    data_portfolio = Reference(ws, min_col=7, min_row=5, max_row=46)
    data_fire = Reference(ws, min_col=8, min_row=5, max_row=46)
    cats = Reference(ws, min_col=3, min_row=6, max_row=46)

    chart.add_data(data_portfolio, titles_from_data=True)
    chart.add_data(data_fire, titles_from_data=True)
    chart.set_categories(cats)

    chart.series[0].graphicalProperties.line.solidFill = EMERALD
    chart.series[0].graphicalProperties.line.width = 28000
    chart.series[1].graphicalProperties.line.solidFill = RED
    chart.series[1].graphicalProperties.line.width = 20000
    chart.series[1].graphicalProperties.line.dashStyle = "dash"

    chart.width = 24
    chart.height = 14
    ws.add_chart(chart, "B48")


# ============================================================
# SHEET 4: INSTRUCTIONS
# ============================================================
def build_instructions(wb):
    ws = wb.create_sheet("How To Use")
    ws.sheet_properties.tabColor = DARK_GRAY
    cols(ws, {"A": 3, "B": 90})

    ws.merge_cells("A1:B2")
    c = ws.cell(row=1, column=1, value="HOW TO USE THE FIRE CALCULATOR PRO")
    c.font = FONT_TITLE
    c.fill = FILL_DARK
    c.alignment = ALIGN_C
    for r in range(1, 3):
        for co in range(1, 3):
            ws.cell(row=r, column=co).fill = FILL_DARK

    sections = [
        ("QUICK START", [
            "1. Open the 'FIRE Calculator' tab and enter your details in the GREEN cells",
            "2. Your FIRE Number appears instantly on the right — this is your target portfolio",
            "3. Check if you're on track: the Gap/Surplus tells you how close you are",
            "4. Use the 'Scenario Comparison' tab to compare different strategies",
            "5. Review the 'Year-by-Year' tab for a detailed projection of your journey",
        ]),
        ("WHAT IS FIRE?", [
            "FIRE = Financial Independence, Retire Early",
            "The goal: build enough investments that your portfolio's returns cover your living expenses",
            "Your FIRE Number = Annual Expenses ÷ Safe Withdrawal Rate",
            "Example: $48,000/year expenses ÷ 4% SWR = $1,200,000 FIRE Number",
            "Once your portfolio hits this number, work becomes OPTIONAL",
        ]),
        ("THE 4% RULE", [
            "Based on the Trinity Study (1998): a portfolio of 50-75% stocks historically survived 30 years",
            "You withdraw 4% of your portfolio in year 1, then adjust for inflation each year",
            "3.5% is more conservative (better for early retirees with 40-50 year horizons)",
            "Variable withdrawal: adjust based on market performance (spend less in downturns)",
            "Guardrails: set upper/lower withdrawal limits to prevent over/under-spending",
        ]),
        ("COAST FIRE vs BARISTA FIRE", [
            "Coast FIRE: You've saved enough that compound growth alone will reach your FIRE number by retirement age",
            "  → You can stop saving and just cover current expenses with any income",
            "Barista FIRE: You've saved enough that part-time income covers the gap",
            "  → Work a low-stress job for benefits + spending money while investments grow",
            "Both are milestones on the path to full FIRE — celebrate them!",
        ]),
        ("SAVINGS RATE: THE #1 LEVER", [
            "Your savings rate determines your timeline more than investment returns",
            "10% savings rate → ~51 years to FI",
            "25% savings rate → ~32 years to FI",
            "50% savings rate → ~17 years to FI",
            "70% savings rate → ~8 years to FI",
            "Focus on: increase income + reduce expenses = higher savings rate = earlier FIRE",
        ]),
        ("INVESTMENT ALLOCATION", [
            "Stocks (equities): Higher return (~10% nominal), higher volatility",
            "Bonds: Lower return (~5% nominal), lower volatility, portfolio stabilizer",
            "Aggressive (90/10): Best for long time horizons (15+ years), higher expected return",
            "Moderate (70/30): Balanced growth and stability, most common FIRE allocation",
            "Conservative (50/50): Lower volatility, suitable for near-retirement or risk-averse",
        ]),
        ("SCENARIO COMPARISON TIPS", [
            "Scenario A automatically links to your main FIRE Calculator inputs",
            "Scenario B: Try aggressive — lower expenses, higher investment, higher stock allocation",
            "Scenario C: Try conservative — more comfortable spending, lower risk",
            "Compare the Gap/Surplus and FIRE ages to find your sweet spot",
            "There's no single 'right' answer — it's about trade-offs between lifestyle now vs. freedom later",
        ]),
        ("COMMON FIRE MILESTONES", [
            "$100K — The hardest milestone (first one). After this, compounding kicks in.",
            "$250K — You're generating meaningful investment income (~$10K/year at 4%)",
            "$500K — Half a million. You can likely Barista FIRE.",
            "Coast FIRE — Your savings can compound to FIRE without adding more. Work is for spending money only.",
            "FIRE Number — Congratulations. Work is now optional.",
        ]),
        ("ABOUT THIS TEMPLATE", [
            "Version: 1.0 | Compatible with: Microsoft Excel 2016+, Google Sheets, LibreOffice Calc",
            "All projections use real (inflation-adjusted) returns for accurate purchasing power comparison",
            "This template is for personal financial planning only.",
            "It is NOT financial advice. Consult a financial advisor for personalized guidance.",
            "© 2026 ClearMetric. All Rights Reserved.",
            "Questions? Visit clearmetric.gumroad.com",
        ]),
    ]

    r = 4
    for title, items in sections:
        ws.cell(row=r, column=2, value=title).font = Font(
            name="Calibri", size=12, bold=True, color=EMERALD)
        ws.cell(row=r, column=2).fill = FILL_LIGHT
        ws.cell(row=r, column=2).border = THIN
        r += 1
        for item in items:
            ws.cell(row=r, column=2, value=item).font = Font(
                name="Calibri", size=11, color="2C3E50")
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 22
            r += 1
        r += 1


# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.Workbook()
    ws = wb.active

    print("Building FIRE Calculator sheet...")
    build_fire_calculator(ws)

    print("Building Scenario Comparison sheet...")
    build_scenario_comparison(wb)

    print("Building Year-by-Year Projection sheet...")
    build_yearly_projection(wb)

    print("Building Instructions sheet...")
    build_instructions(wb)

    wb.active = 0

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "output", "ClearMetric-FIRE-Calculator-Pro.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"Size: {os.path.getsize(out) / 1024:.1f} KB")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
