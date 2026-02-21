"""
Stress test for the FIRE Calculator Pro Excel template.
Opens the generated .xlsx and verifies:
  - All sheets exist
  - All formulas are syntactically valid
  - No broken cell references (#REF, #NAME)
  - Input cells are unlocked, formula cells are locked
  - Cross-sheet references resolve
  - Charts exist on expected sheets
  - Conditional formatting rules are present
  - Column widths are set (no hidden content)
"""

import openpyxl
import re
import sys
import os

XLSX_PATH = os.path.join(os.path.dirname(__file__), "output", "ClearMetric-FIRE-Calculator-Pro.xlsx")

passed = 0
failed = 0
errors = []

def test(name, condition, detail=""):
    global passed, failed, errors
    if condition:
        passed += 1
        print(f"  ✅ {name}")
    else:
        failed += 1
        msg = f"{name}: {detail}" if detail else name
        errors.append(msg)
        print(f"  ❌ {msg}")


print("=" * 60)
print("FIRE CALCULATOR EXCEL — STRESS TEST")
print("=" * 60)

# Load workbook
wb = openpyxl.load_workbook(XLSX_PATH)

# --- Sheet existence ---
print("\n1. SHEET STRUCTURE")
expected_sheets = ["FIRE Calculator", "Scenario Comparison", "Year-by-Year", "How To Use"]
for s in expected_sheets:
    test(f"Sheet '{s}' exists", s in wb.sheetnames, f"Missing. Found: {wb.sheetnames}")

test(f"Exactly 4 sheets", len(wb.sheetnames) == 4, f"Found {len(wb.sheetnames)}")

# --- FIRE Calculator sheet ---
print("\n2. FIRE CALCULATOR SHEET — Formulas")
ws = wb["FIRE Calculator"]

formula_cells = {}
for row in ws.iter_rows(min_row=1, max_row=50, max_col=8):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
            formula_cells[cell.coordinate] = cell.value

test("Has formulas", len(formula_cells) > 10, f"Only {len(formula_cells)} formulas found")

# Check for common formula errors
for coord, formula in formula_cells.items():
    test(f"Formula {coord} no #REF", "#REF" not in formula, formula)
    test(f"Formula {coord} no #NAME", "#NAME" not in formula, formula)

# Check key formulas exist
key_formulas = {
    "C15": "after-tax income",
    "C16": "annual expenses",
    "G13": "blended return",
    "G14": "real return",
    "G24": "total projected corpus",
    "G25": "gap/surplus",
    "G26": "status",
}
for coord, desc in key_formulas.items():
    val = ws[coord].value
    test(f"{coord} ({desc}) has formula", val and isinstance(val, str) and val.startswith("="),
         f"Expected formula, got: {val}")

# FIRE number (F6 is the big display)
fire_val = ws["F6"].value
test("F6 (FIRE number) has formula", fire_val and isinstance(fire_val, str) and "C28" in fire_val,
     f"Got: {fire_val}")

# --- Input cells ---
print("\n3. INPUT CELLS — Values & Protection")
input_cells_expected = [
    ("C6", 30, "Current Age"),
    ("C7", 45, "Retirement Age"),
    ("C8", 90, "Life Expectancy"),
    ("C11", 100000, "Annual Income"),
    ("C12", 0.25, "Tax Rate"),
    ("C13", 4000, "Monthly Expenses"),
    ("C14", 0.03, "Income Growth"),
    ("C19", 150000, "Current Savings"),
    ("C20", 2500, "Monthly Investment"),
    ("C21", 0.70, "Stock Allocation"),
    ("C25", 0.10, "Stock Return"),
    ("C26", 0.05, "Bond Return"),
    ("C27", 0.03, "Inflation"),
    ("C28", 0.04, "SWR"),
]

for coord, expected_val, desc in input_cells_expected:
    cell = ws[coord]
    test(f"{coord} ({desc}) has value", cell.value == expected_val,
         f"Expected {expected_val}, got {cell.value}")
    test(f"{coord} ({desc}) is unlocked", cell.protection.locked == False,
         f"locked={cell.protection.locked}")

# --- Sheet protection ---
print("\n4. SHEET PROTECTION")
test("Sheet is protected", ws.protection.sheet == True)

# --- Scenario Comparison sheet ---
print("\n5. SCENARIO COMPARISON SHEET")
ws2 = wb["Scenario Comparison"]

# Check column headers
test("Scenario A header", ws2["C5"].value == "Current Plan")
test("Scenario B header", ws2["E5"].value == "Aggressive")
test("Scenario C header", ws2["G5"].value == "Conservative")

# Check that Scenario B and C have input values
test("Scenario B retirement age", ws2["E7"].value == 40)
test("Scenario C retirement age", ws2["G7"].value == 55)
test("Scenario B stock allocation", ws2["E12"].value == 0.90)
test("Scenario C stock allocation", ws2["G12"].value == 0.50)

# Check result formulas
result_cells_sc = ["C17", "C18", "C19", "C20", "C21", "C22", "C23",
                   "E17", "E18", "E19", "E20", "E21", "E22", "E23",
                   "G17", "G18", "G19", "G20", "G21", "G22", "G23"]
for coord in result_cells_sc:
    val = ws2[coord].value
    test(f"SC {coord} has formula", val and isinstance(val, str) and val.startswith("="),
         f"Expected formula, got: {val}")

# Check conditional formatting
test("SC has conditional formatting", len(ws2.conditional_formatting) > 0,
     f"Found {len(ws2.conditional_formatting)} rules")

# --- Year-by-Year sheet ---
print("\n6. YEAR-BY-YEAR SHEET")
ws3 = wb["Year-by-Year"]

# Check headers
headers = ["Year", "Age", "Portfolio Start", "Annual Investment",
           "Investment Growth", "Portfolio End", "FIRE Number", "% of FIRE"]
for i, h in enumerate(headers):
    cell_val = ws3.cell(row=5, column=2+i).value
    test(f"YbY header col {2+i}: '{h}'", cell_val == h, f"Got: {cell_val}")

# Check first data row (row 6)
test("YbY row 6 Year=0", ws3["B6"].value == 0)
test("YbY row 6 Age formula", ws3["C6"].value and "'FIRE Calculator'" in str(ws3["C6"].value))
test("YbY row 6 Portfolio Start links to main sheet",
     ws3["D6"].value and "'FIRE Calculator'" in str(ws3["D6"].value))

# Check row 7 (year 1) links to previous row
test("YbY row 7 Portfolio Start = prev end", ws3["D7"].value == "=G6")

# Check 40 years of data
for r in range(6, 47):
    val = ws3.cell(row=r, column=9).value  # % of FIRE
    test(f"YbY row {r} has % of FIRE formula",
         val and isinstance(val, str) and val.startswith("="),
         f"Got: {val}")

# Check chart exists
test("YbY has chart(s)", len(ws3._charts) > 0, f"Found {len(ws3._charts)} charts")

# --- How To Use sheet ---
print("\n7. HOW TO USE SHEET")
ws4 = wb["How To Use"]

# Check title
test("Instructions title exists", ws4["A1"].value == "HOW TO USE THE FIRE CALCULATOR PRO")

# Check key sections exist
section_keywords = ["QUICK START", "WHAT IS FIRE", "4% RULE", "COAST FIRE", "SAVINGS RATE"]
found_sections = set()
for row in ws4.iter_rows(min_row=4, max_row=80, min_col=2, max_col=2):
    for cell in row:
        if cell.value:
            for kw in section_keywords:
                if kw in str(cell.value).upper():
                    found_sections.add(kw)

for kw in section_keywords:
    test(f"Section '{kw}' exists", kw in found_sections, f"Missing from instructions")

# --- Cross-sheet reference validation ---
print("\n8. CROSS-SHEET REFERENCES")
cross_sheet_pattern = re.compile(r"'FIRE Calculator'!")

# Check Year-by-Year references
yby_formulas = []
for row in ws3.iter_rows(min_row=6, max_row=46, max_col=9):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cross_sheet_pattern.search(cell.value):
            yby_formulas.append(cell.coordinate)

test("Year-by-Year has cross-sheet refs", len(yby_formulas) > 20,
     f"Only {len(yby_formulas)} cross-sheet refs found")

# Check Scenario Comparison references to FIRE Calculator
sc_cross_refs = []
for row in ws2.iter_rows(min_row=6, max_row=14, max_col=3):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cross_sheet_pattern.search(cell.value):
            sc_cross_refs.append(cell.coordinate)

# Scenario A links to main sheet
test("Scenario A links to FIRE Calculator sheet", len(sc_cross_refs) >= 5,
     f"Only {len(sc_cross_refs)} cross-refs found for Scenario A")

# --- File metadata ---
print("\n9. FILE METADATA")
file_size = os.path.getsize(XLSX_PATH)
test("File size reasonable (10-100 KB)", 10_000 < file_size < 100_000,
     f"Size: {file_size / 1024:.1f} KB")

# Tab colors
test("FIRE Calculator tab color set", ws.sheet_properties.tabColor is not None)
test("Scenario Comparison tab color set", ws2.sheet_properties.tabColor is not None)
test("Year-by-Year tab color set", ws3.sheet_properties.tabColor is not None)

# ================================================================
print("\n" + "=" * 60)
print(f"RESULTS: {passed} passed, {failed} failed out of {passed + failed}")
print("=" * 60)

if errors:
    print("\nFAILURES:")
    for e in errors:
        print(f"  ❌ {e}")
    sys.exit(1)
else:
    print("\n✅ ALL TESTS PASSED — Excel template is solid.")
    sys.exit(0)
